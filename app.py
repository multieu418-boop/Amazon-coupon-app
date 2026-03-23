import streamlit as st
import pandas as pd
import io
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta

# --- 页面设置 ---
st.set_page_config(page_title="Cupshe 亚马逊优惠券助手", layout="wide")

# --- 自定义样式 ---
st.markdown("""
    <style>
    .main { background-color: #f5f7f9; }
    .stButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #ff4b4b; color: white; }
    .stDownloadButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #008cba; color: white; }
    </style>
    """, unsafe_allow_html=True)

st.title("👗 Cupshe 亚马逊优惠券智能管理工具")
st.info("集初始创建与报错纠错于一体的闭环工作流")

# --- 侧边栏：基础数据上传 ---
st.sidebar.header("⚙️ 核心数据配置")
inventory_file = st.sidebar.file_uploader("第一步：上传 DE.txt (All Listing Report)", type=['txt'])

@st.cache_data
def load_inventory(file):
    if file:
        try:
            # 自动处理编码问题
            df = pd.read_csv(file, sep='\t', encoding='utf-8')
        except:
            df = pd.read_csv(file, sep='\t', encoding='cp1252')
        return df[['asin1', 'price']].drop_duplicates('asin1').set_index('asin1')
    return None

inv_data = load_inventory(inventory_file)

if inv_data is not None:
    st.sidebar.success(f"✅ 已加载 {len(inv_data)} 个 ASIN 价格数据")
else:
    st.sidebar.warning("⚠️ 请先上传 DE.txt 激活价格校验功能")

# --- 标签页切换 ---
tab1, tab2 = st.tabs(["第一阶段：创建新券", "第二阶段：纠错修复"])

# =================================================================
# 视图 1：第一阶段 - 初始创建
# =================================================================
with tab1:
    st.header("1️⃣ 初始创建优惠券模板")
    
    with st.form("create_form"):
        col1, col2 = st.columns(2)
        with col1:
            asin_input = st.text_area("输入子 ASIN 列表 (分号/逗号/换行分隔)", height=150, placeholder="B0xxxxxx;B0yyyyyy...")
            disc_type = st.selectbox("折扣类型", ["折扣 (Percentage)", "满减 (Money)"])
            disc_val = st.number_input("折扣数额", min_value=1.0, value=20.0, step=1.0)
            
        with col2:
            coupon_suffix = st.text_input("优惠券名称后缀", value="Cupshe Summer Sale")
            budget = st.number_input("预算 (€)", min_value=100, value=1000)
            c_dates = st.date_input("优惠券日期范围", [datetime.now() + timedelta(days=1), datetime.now() + timedelta(days=30)])
            
        submitted = st.form_submit_button("生成上传模板")
        
        if submitted:
            if not asin_input:
                st.error("请输入 ASIN！")
            else:
                # 清洗 ASIN
                asins = [a.strip() for a in re.split(r'[;\n,\s]+', asin_input) if a.strip()]
                
                # 价格预检
                if inv_data is not None:
                    invalid_asins = []
                    for a in asins:
                        if a in inv_data.index:
                            p = inv_data.loc[a, 'price']
                            if disc_type == "满减 (Money)" and disc_val >= p * 0.5:
                                invalid_asins.append(f"{a}(价€{p})")
                    if invalid_asins:
                        st.warning(f"⚠️ 预警：以下 ASIN 折扣可能过大(超50%): {', '.join(invalid_asins)}")

                # 构建 Amazon 模板 (对应你文件的第7行逻辑)
                # 提示：Amazon 模板前几行通常是说明，这里为了简化直接输出数据行
                data = {
                    "ASIN 列表": ";".join(asins),
                    "折扣类型": "Percentage" if "折扣" in disc_type else "Money",
                    "折扣数额": disc_val,
                    "每张减免金额": disc_val if "满减" in disc_type else "",
                    "名称": coupon_suffix,
                    "预算": budget,
                    "开始日期": c_dates[0].strftime("%Y-%m-%d"),
                    "结束日期": c_dates[1].strftime("%Y-%m-%d") if len(c_dates)>1 else "",
                    "限购一次": "Yes",
                    "目标客户": "All Customers"
                }
                df_out = pd.DataFrame([data])
                
                output = io.BytesIO()
                df_out.to_excel(output, index=False)
                st.download_button("💾 下载第一阶段上传文件", output.getvalue(), "Phase1_Upload.xlsx")

# =================================================================
# 视图 2：第二阶段 - 纠错修复
# =================================================================
with tab2:
    st.header("2️⃣ 报错纠错与自动重做")
    st.write("上传亚马逊返回的带批注的报错文件，系统将自动解析。")
    
    err_file = st.file_uploader("上传报错的 Excel 文件 (.xlsx)", type=['xlsx'])
    
    if err_file:
        # 使用 openpyxl 解析批注
        wb = load_workbook(err_file)
        ws = wb.active
        
        error_records = []
        # 假设数据从第8行开始 (对应亚马逊模板)
        for row_idx in range(8, ws.max_row + 1):
            n_cell = ws.cell(row=row_idx, column=14) # N列
            # 即使没有 Comment，如果单元格内容包含错误提示也抓取
            if n_cell.comment or (n_cell.value and "验证" in str(n_cell.value)):
                comment_text = n_cell.comment.text if n_cell.comment else str(n_cell.value)
                raw_asin_str = str(ws.cell(row=row_idx, column=1).value)
                
                # 提取批注中的具体 ASIN
                bug_asins = re.findall(r'[A-Z0-9]{10}', comment_text)
                all_asins = raw_asin_str.split(';')
                
                error_records.append({
                    "row": row_idx,
                    "coupon_name": ws.cell(row=row_idx, column=5).value,
                    "bug_asins": list(set(bug_asins)),
                    "all_asins": all_asins,
                    "orig_row_data": [ws.cell(row=row_idx, column=c).value for c in range(1, 15)]
                })

        if not error_records:
            st.success("未在文件中检测到 N 列报错批注，请确认文件是否正确。")
        else:
            st.write(f"🔍 检测到 {len(error_records)} 组优惠券存在报错。")
            
            final_rows = []
            for rec in error_records:
                st.markdown(f"---")
                st.subheader(f"优惠券: {rec['coupon_name']}")
                
                # 让用户针对这组报错 ASIN 做决策
                col_a, col_b = st.columns(2)
                with col_a:
                    st.write(f"❌ 报错 ASIN: `{', '.join(rec['bug_asins'])}`")
                    action = st.radio("决策策略:", ["全部剔除", "手动修正报错 ASIN"], key=f"radio_{rec['row']}")
                
                # 准备剩下的 ASIN
                clean_asins = [a for a in rec['all_asins'] if a not in rec['bug_asins']]
                
                if action == "全部剔除":
                    # 只保留干净的 ASIN
                    new_row = list(rec['orig_row_data'])
                    new_row[0] = ";".join(clean_asins)
                    final_rows.append(new_row)
                    st.caption("✅ 已从原计划中移除报错 ASIN")
                
                else:
                    # 修正逻辑
                    new_val = st.number_input(f"为报错 ASIN 设置新力度 (建议参考 DE.txt)", min_value=1.0, value=5.0, key=f"val_{rec['row']}")
                    # 1. 保留原行（剔除报错部分）
                    row_remain = list(rec['orig_row_data'])
                    row_remain[0] = ";".join(clean_asins)
                    final_rows.append(row_remain)
                    # 2. 新增一行（专门跑这几个报错的）
                    row_fix = list(rec['orig_row_data'])
                    row_fix[0] = ";".join(rec['bug_asins'])
                    row_fix[2] = new_val # 修改力度
                    row_fix[4] = f"{rec['coupon_name']}-FIX" # 修改名称区分
                    final_rows.append(row_fix)
                    st.caption("✅ 已将报错 ASIN 拆分为新 Coupon 行")

            if st.button("🚀 合并并生成最终纠错文件"):
                final_df = pd.DataFrame(final_rows)
                # 导出
                out_fix = io.BytesIO()
                final_df.to_excel(out_fix, index=False, header=False)
                st.download_button("💾 下载最终修复版模板", out_fix.getvalue(), "Final_Fixed_Coupons.xlsx")